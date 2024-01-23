import csv
from datetime import datetime
import os
import logging
import openpyxl
from collections import Counter
from datetime import datetime

PATH_TO_CSV = './logs/user_login.csv'
PATH_TO_XLSX = './logs/login_history.xlsx'


def log_user_login(username):
    logging.info(f'{username}')
    write_to_excel(username)


def write_to_excel(username):
    if os.path.exists(PATH_TO_XLSX):
        workbook = openpyxl.load_workbook(PATH_TO_XLSX)
        sheet = workbook.active

        timestamp = datetime.now().strftime('%Y-%m-%d %H')
        current_timestamp_username = f'{timestamp} {username}'

        need_to_update_count = False
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            excel_row_string = f"{row[0]}-{row[1]:02}-{row[2]:02} {row[3][:2]} {row[4]}"

            if current_timestamp_username == excel_row_string:
                need_to_update_count = True
                current_count_cell = sheet.cell(row=row_number, column=6)
                current_count_cell.value += 1
                break
        if not need_to_update_count:
            date_part, hour_part = timestamp.split(' ')
            year, month, day = date_part.split('-')
            row = [int(year), int(month), int(day), f'{int(hour_part)}:00 - {int(hour_part)}:59', username, 1]
            sheet.append(row)
        workbook.save(PATH_TO_XLSX)
    else:
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            column_names = ['Year', 'Month',
                            'Day', 'Hour', 'Username', 'Count']
            sheet.append(column_names)

            with open(PATH_TO_CSV, 'r', newline='', encoding='utf-8') as file:
                csv_data = list(csv.reader(file))

            counts_per_hour = Counter()
            for username, timestamp in csv_data:
                dt = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                timestamp_slice = dt.strftime('%Y-%m-%d %H')
                counts_per_hour[(username, timestamp_slice)] += 1

            for (username, timestamp_slice), count in counts_per_hour.items():
                date_part, hour_part = timestamp_slice.split(' ')
                year, month, day = date_part.split('-')
                row = [int(year), int(month), int(day), f'{int(hour_part)}:00 - {int(hour_part)}:59', username, count]
                sheet.append(row)

            workbook.save(PATH_TO_XLSX)
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
